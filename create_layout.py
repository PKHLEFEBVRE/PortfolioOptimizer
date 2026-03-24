import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def create_excel_layout(filename="Investment_Simulation.xlsx"):
    wb = openpyxl.Workbook()

    # Dashboard Sheet
    ws_dashboard = wb.active
    ws_dashboard.title = "Dashboard"

    # Headers for Dashboard
    ws_dashboard["A1"] = "Investment Simulation Summary"
    ws_dashboard["A1"].font = Font(size=14, bold=True)

    ws_dashboard["A3"] = "Total Profit Investor 1 (All-in):"
    ws_dashboard["A4"] = "Total Profit Investor 2 (Scale-in):"

    # Security Breakdown Headers
    headers_dashboard = [
        "Security", "Drift", "Volatility",
        "Inv 1 Profit", "Inv 2 Profit",
        "Inv 1 Trade Day", "Inv 1 Trade Price",
        "Inv 2 Trade 1 Day", "Inv 2 Trade 1 Price",
        "Inv 2 Trade 2 Day", "Inv 2 Trade 2 Price",
        "Inv 2 Trade 3 Day", "Inv 2 Trade 3 Price"
    ]

    for col_num, header in enumerate(headers_dashboard, 1):
        cell = ws_dashboard.cell(row=6, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Adjust column widths
    for col in ws_dashboard.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_dashboard.column_dimensions[column].width = adjusted_width

    # Simulation Data Sheet
    ws_data = wb.create_sheet(title="Simulation Data")

    headers_data = ["Day"] + [f"Security {i}" for i in range(1, 11)]
    for col_num, header in enumerate(headers_data, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws_data.column_dimensions[cell.column_letter].width = 12

    # Save workbook
    wb.save(filename)
    print(f"Created {filename}")

if __name__ == "__main__":
    create_excel_layout()
